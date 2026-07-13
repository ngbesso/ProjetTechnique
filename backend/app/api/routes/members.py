import csv
import io
import secrets
from datetime import date, datetime
from typing import Annotated

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.email import (
    EmailSender,
    get_email_sender,
    membership_approved,
    membership_approved_invite,
    membership_received,
)
from app.core.config import settings
from app.core.security import create_setup_token, hash_password
from app.db.session import get_db
from app.models.church import Church
from app.models.member import Member, MemberStatus
from app.models.rbac import Role, UserRole
from app.models.setting import AppSetting
from app.models.user import User
from app.schemas.member import (
    MemberCreate,
    MemberImportResult,
    MemberImportRowError,
    MemberList,
    MemberRead,
    MemberSelfUpdate,
    MembershipRequest,
    MemberUpdate,
)

_IMPORT_REQUIRED_COLUMNS = {"first_name", "last_name", "email"}
_IMPORT_COLUMNS = (
    "first_name",
    "last_name",
    "email",
    "address",
    "birth_date",
    "sexe",
    "telephone",
    "family_status",
    "conversion_date",
    "is_baptized",
)
_IMPORT_EXAMPLE_ROW = (
    "Jean",
    "Dupont",
    "jean.dupont@exemple.com",
    "123 Rue Principale",
    "1985-06-14",
    "Masculin",
    "5145551234",
    "Marie(e)",
    "2010-09-01",
    "oui",
)


def _parse_bool(value: str | None) -> bool:
    return (value or "").strip().lower() in ("1", "true", "vrai", "oui", "yes", "y")


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "oui" if value else "non"
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), [dict(row) for row in reader]


def _rows_from_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, ())
    fieldnames = [str(h).strip() if h is not None else "" for h in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append(
            {
                fieldnames[i]: _cell_to_str(values[i])
                for i in range(len(fieldnames))
                if i < len(values)
            }
        )
    return fieldnames, rows


def _build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"
    ws.append(_IMPORT_COLUMNS)
    ws.append(_IMPORT_EXAMPLE_ROW)
    for i, header in enumerate(_IMPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 14)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


router = APIRouter(prefix="/members", tags=["membres"])


def _load(db: Session, member_id: int) -> Member:
    member = db.get(Member, member_id)
    if not member:
        raise HTTPException(404, "Membre introuvable")
    return member


def _ensure(user: User, member: Member, code: str) -> None:
    if not user.has_permission(code, member.church_id):
        raise HTTPException(403, "Permission insuffisante sur cette église")


_EMAIL_TAKEN = "Cette adresse courriel ne peut pas être utilisée. Veuillez en choisir une autre ou contacter l'administrateur si vous pensez qu'il s'agit d'une erreur."


def _check_email_unique(db: Session, email: str, exclude_id: int | None = None) -> None:
    """Lève HTTP 409 avec un message générique si l'email est déjà pris."""
    query = select(Member).where(Member.email == email)
    if exclude_id is not None:
        query = query.where(Member.id != exclude_id)
    if db.scalar(query):
        raise HTTPException(409, _EMAIL_TAKEN)
    if db.scalar(select(User).where(User.email == email)):
        raise HTTPException(409, _EMAIL_TAKEN)


def _generate_member_code(db: Session) -> str:
    year = date.today().year
    prefix = f"MBR-{year}-"
    count = (
        db.scalar(
            select(func.count())
            .select_from(Member)
            .where(Member.member_code.like(f"{prefix}%"))
        )
        or 0
    )
    return f"{prefix}{count + 1:04d}"


def _auto_approve_enabled(db: Session) -> bool:
    row = db.get(AppSetting, "auto_approve_members")
    return row is not None and row.value == "true"


def _do_approve(
    member: Member,
    db: Session,
    background: BackgroundTasks,
    sender: EmailSender,
) -> None:
    """Approuve un membre : active le compte, crée/lie l'utilisateur, envoie l'email."""
    member.status = MemberStatus.active
    if not member.member_code:
        member.member_code = _generate_member_code(db)

    user = db.scalar(select(User).where(User.email == member.email))
    invite_link: str | None = None
    if user is None:
        user = User(
            email=member.email,
            hashed_password=hash_password(secrets.token_urlsafe(16)),
        )
        db.add(user)
        db.flush()
        token = create_setup_token(user.id, user.token_version)
        invite_link = f"{settings.frontend_url}/definir-mot-de-passe?token={token}"
    member.user_id = user.id

    role_membre = db.scalar(select(Role).where(Role.name == "membre"))
    if role_membre:
        exists = db.scalar(
            select(UserRole).where(
                UserRole.user_id == user.id,
                UserRole.role_id == role_membre.id,
                UserRole.church_id == member.church_id,
            )
        )
        if not exists:
            db.add(
                UserRole(
                    user_id=user.id, role_id=role_membre.id, church_id=member.church_id
                )
            )

    if invite_link:
        background.add_task(
            membership_approved_invite,
            sender,
            member.email,
            member.first_name,
            invite_link,
        )
    else:
        background.add_task(
            membership_approved, sender, member.email, member.first_name
        )


@router.post("/request", response_model=MemberRead, status_code=201)
def request_membership(
    data: MembershipRequest,
    background: BackgroundTasks,
    db: Annotated[Session, Depends(get_db)],
    sender: Annotated[EmailSender, Depends(get_email_sender)],
):
    if not db.get(Church, data.church_id):
        raise HTTPException(404, "Église introuvable")
    _check_email_unique(db, data.email)

    member = Member(**data.model_dump(), status=MemberStatus.pending)
    db.add(member)
    db.flush()

    if _auto_approve_enabled(db):
        _do_approve(member, db, background, sender)
    else:
        background.add_task(
            membership_received, sender, member.email, member.first_name
        )

    db.commit()
    db.refresh(member)
    return member


@router.get("/me", response_model=MemberRead)
def my_profile(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = db.scalar(select(Member).where(Member.user_id == current_user.id))
    if not member:
        raise HTTPException(404, "Aucun profil de membre associé")
    return member


@router.patch("/me", response_model=MemberRead)
def update_my_profile(
    data: MemberSelfUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = db.scalar(select(Member).where(Member.user_id == current_user.id))
    if not member:
        raise HTTPException(404, "Aucune fiche membre liée à ce compte")

    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(member, k, v)
    db.commit()
    db.refresh(member)
    return member


@router.get("", response_model=MemberList)
def list_members(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    q: str | None = None,
    status: MemberStatus | None = None,
    limit: int = Query(default=20, le=100),
    offset: int = 0,
):
    scope = current_user.accessible_church_ids("member:read")
    if scope is not None and not scope:
        raise HTTPException(403, "Aucun périmètre accessible")
    query = select(Member)
    if scope is not None:
        query = query.where(Member.church_id.in_(scope))
    if q:
        like = f"%{q}%"
        query = query.where(
            or_(
                Member.first_name.ilike(like),
                Member.last_name.ilike(like),
                Member.email.ilike(like),
            )
        )
    if status:
        query = query.where(Member.status == status)
    total = db.scalar(select(func.count()).select_from(query.subquery()))
    rows = db.scalars(
        query.order_by(Member.created_at.desc()).limit(limit).offset(offset)
    ).all()
    return MemberList(
        items=[MemberRead.model_validate(m) for m in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.post("", response_model=MemberRead, status_code=201)
def create_member(
    data: MemberCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    if not current_user.has_permission("member:create", data.church_id):
        raise HTTPException(403, "Permission insuffisante sur cette église")
    if not db.get(Church, data.church_id):
        raise HTTPException(404, "Église introuvable")
    _check_email_unique(db, data.email)
    member = Member(**data.model_dump(), status=MemberStatus.active)
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


@router.get("/import/template")
def download_import_template(
    current_user: Annotated[User, Depends(get_current_user)],
):
    """Modèle Excel (colonnes attendues + exemple) pour l'import en masse de membres."""
    scope = current_user.accessible_church_ids("member:create")
    if scope is not None and not scope:
        raise HTTPException(403, "Permission insuffisante")
    content = _build_import_template()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=modele-import-membres.xlsx"
        },
    )


@router.post("/import", response_model=MemberImportResult)
def import_members(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    church_id: Annotated[int, Form()],
    file: Annotated[UploadFile, File()],
):
    """Importe en masse des membres déjà connus d'une église (statut actif direct). Accepte .csv et .xlsx."""
    if not current_user.has_permission("member:create", church_id):
        raise HTTPException(403, "Permission insuffisante sur cette église")
    if not db.get(Church, church_id):
        raise HTTPException(404, "Église introuvable")

    content = file.file.read()
    if (file.filename or "").lower().endswith(".xlsx"):
        fieldnames, rows = _rows_from_xlsx(content)
    else:
        fieldnames, rows = _rows_from_csv(content)

    missing_columns = _IMPORT_REQUIRED_COLUMNS - set(fieldnames)
    if missing_columns:
        raise HTTPException(
            422,
            f"Colonnes manquantes dans le fichier : {', '.join(sorted(missing_columns))}",
        )

    errors: list[MemberImportRowError] = []
    seen_emails: set[str] = set()
    created = 0

    for i, row in enumerate(rows, start=2):  # ligne 1 = en-têtes
        email = (row.get("email") or "").strip().lower()
        try:
            data = MemberCreate.model_validate(
                {
                    "church_id": church_id,
                    "first_name": (row.get("first_name") or "").strip(),
                    "last_name": (row.get("last_name") or "").strip(),
                    "email": email,
                    "address": (row.get("address") or "").strip() or None,
                    "birth_date": (row.get("birth_date") or "").strip() or None,
                    "sexe": (row.get("sexe") or "").strip() or None,
                    "telephone": (row.get("telephone") or "").strip() or None,
                    "family_status": (row.get("family_status") or "").strip() or None,
                    "conversion_date": (row.get("conversion_date") or "").strip()
                    or None,
                    "is_baptized": _parse_bool(row.get("is_baptized")),
                }
            )
        except ValidationError as exc:
            errors.append(
                MemberImportRowError(
                    row=i, email=email or None, message=exc.errors()[0]["msg"]
                )
            )
            continue

        if email in seen_emails:
            errors.append(
                MemberImportRowError(
                    row=i, email=email, message="Email en double dans le fichier."
                )
            )
            continue
        try:
            _check_email_unique(db, data.email)
        except HTTPException:
            errors.append(
                MemberImportRowError(
                    row=i,
                    email=email,
                    message="Cette adresse courriel est déjà utilisée.",
                )
            )
            continue

        member = Member(
            **data.model_dump(),
            status=MemberStatus.active,
            member_code=_generate_member_code(db),
        )
        db.add(member)
        db.flush()
        seen_emails.add(email)
        created += 1

    db.commit()
    return MemberImportResult(created=created, errors=errors)


@router.get("/{member_id}", response_model=MemberRead)
def get_member(
    member_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:read")
    return member


@router.patch("/{member_id}", response_model=MemberRead)
def update_member(
    member_id: int,
    data: MemberUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:update")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(member, k, v)
    db.commit()
    db.refresh(member)
    return member


@router.post("/{member_id}/approve", response_model=MemberRead)
def approve_member(
    member_id: int,
    background: BackgroundTasks,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    sender: Annotated[EmailSender, Depends(get_email_sender)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:approve")
    _do_approve(member, db, background, sender)
    db.commit()
    db.refresh(member)
    return member


@router.post("/{member_id}/reject", response_model=MemberRead)
def reject_member(
    member_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:approve")
    member.status = MemberStatus.rejected
    db.commit()
    db.refresh(member)
    return member


@router.post("/{member_id}/deactivate", response_model=MemberRead)
def deactivate_member(
    member_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:update")
    member.status = MemberStatus.inactive
    db.commit()
    db.refresh(member)
    return member


@router.post("/{member_id}/activate", response_model=MemberRead)
def activate_member(
    member_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    member = _load(db, member_id)
    _ensure(current_user, member, "member:update")
    member.status = MemberStatus.active
    db.commit()
    db.refresh(member)
    return member
