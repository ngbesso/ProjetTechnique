import io

import openpyxl
import pytest

from app.services import report_builder


@pytest.fixture(autouse=True)
def _no_ai_summary(monkeypatch):
    """Évite un appel réseau réel vers ai-service pendant les tests."""
    monkeypatch.setattr(report_builder, "fetch_ai_summary", lambda *a, **k: None)


def test_requires_auth(client):
    r = client.get("/admin/reports/eglises?format=excel")
    assert r.status_code == 401


def test_requires_admin(client, make_user, auth_header):
    make_user("m@b.com", roles=["membre"])
    r = client.get(
        "/admin/reports/eglises?format=excel", headers=auth_header("m@b.com")
    )
    assert r.status_code == 403


def test_unknown_domain(client, make_user, auth_header):
    make_user("admin@b.com", roles=["admin"])
    r = client.get(
        "/admin/reports/inconnu?format=excel", headers=auth_header("admin@b.com")
    )
    assert r.status_code == 400


def test_unknown_format(client, make_user, auth_header):
    make_user("admin@b.com", roles=["admin"])
    r = client.get(
        "/admin/reports/eglises?format=powerpoint", headers=auth_header("admin@b.com")
    )
    assert r.status_code == 400


def test_excel_report_content(client, make_user, auth_header):
    make_user("admin@b.com", roles=["admin"])
    r = client.get(
        "/admin/reports/eglises?format=excel", headers=auth_header("admin@b.com")
    )
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    values = {
        row[0]: row[1] for row in wb["Résumé"].iter_rows(values_only=True) if row[0]
    }
    assert values["total"] == "1"
    assert "by_district" in wb.sheetnames


def test_word_report_returns_docx(client, make_user, auth_header):
    make_user("admin@b.com", roles=["admin"])
    r = client.get(
        "/admin/reports/sermons?format=word", headers=auth_header("admin@b.com")
    )
    assert r.status_code == 200
    assert "wordprocessingml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"  # signature .docx (zip)


def test_pdf_report_returns_pdf(client, make_user, auth_header):
    make_user("admin@b.com", roles=["admin"])
    r = client.get("/admin/reports/dons?format=pdf", headers=auth_header("admin@b.com"))
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF")
